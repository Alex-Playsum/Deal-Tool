"""Reddit Deal Table Tool - GUI entry point."""

import json
import math
import os
import queue
import random
import re
import sys
import threading
import webbrowser
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog, simpledialog

from config import (
    ALL_CURRENCIES,
    CURRENCY_LABELS,
    DEFAULT_CURRENCIES,
    STEAM_LABEL_MIN_PERCENT,
    STEAM_LABEL_ORDER,
)
from deal_filters import (
    apply_deal_filters,
    apply_discount_filter,
    apply_price_filter,
    date_str_to_start_of_day_ms,
    ONE_DAY_MS,
    parse_sale_end_value,
)
from feed_client import fetch_and_parse
from product_index import items_to_index, normalize_url, resolve_urls_to_products
from table_builder import build_reddit_table
from on_sale import (
    get_on_sale_products,
    enrich_with_steam_reviews,
    _discount_str,
    _discount_pct,
    _sale_end_ms,
    _sale_end_str,
    _release_date_str,
)
from tksheet import Sheet
from steam_cache import clear as clear_steam_cache
from steam_app_list import clear_app_list_cache, clear_name_resolution_cache
from steam_appdetails_cache import clear as clear_steam_appdetails_cache
from steamspy_client import clear_steamspy_cache
from email_html import build_email_html
from steam_client import fetch_app_details_full

# Directory for saved email templates (one JSON file per template)
EMAIL_TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "email_templates")


def _email_build_worker(
    worker_queue: queue.Queue,
    index: dict,
    params: dict,
    blocks: list,
    pre_enriched_rows: list | None = None,
) -> None:
    """Run in background thread: build game pool, enrich, build email HTML. Puts (done|error) on queue.
    When source is auto and pre_enriched_rows is set (in-memory cache from Deal Finder or previous build),
    reuses it instead of re-fetching. When doing a full build, sends full_rows back so main thread can update _on_sale_rows."""
    def put(msg):
        worker_queue.put(msg)
    full_rows_for_cache = None
    try:
        put(("progress", "email", "Getting game pool…"))
        source = (params.get("source") or "auto").strip() or "auto"
        currency = (params.get("currency") or "USD").strip() or "USD"
        if source == "list":
            urls = parse_pasted_urls(params.get("urls_text") or "")
            products, _ = resolve_urls_to_products(index, urls)
            pool = list(products)
        elif source == "auto" and pre_enriched_rows:
            put(("progress", "email", "Using in-memory cache…"))
            rows = [dict(r) for r in pre_enriched_rows]
            currency = (params.get("currency") or "USD").strip() or "USD"
            rows = apply_deal_filters(
                rows,
                score_type=params.get("score_type") or "All",
                score_value=params.get("score_value") or "",
                label_value=params.get("label_value") or "",
                min_reviews=params.get("min_reviews") or "",
                discount_value=params.get("discount_value") or "",
                price_value=params.get("price_value") or "",
                currency=currency,
                sale_end_type=params.get("sale_end_type") or "All",
                sale_end_value=params.get("sale_end_value") or "",
            )
            rows = _apply_publisher_filter(rows, params.get("publisher") or "")
            rows = _apply_developer_filter(rows, params.get("developer") or "")
            rows = _apply_tags_filter(rows, params.get("tags") or "")
            pool = rows
        else:
            products = get_on_sale_products(index, resolve_steam_by_name=True)
            put(("progress", "email", "Fetching Steam data…"))
            def progress(i, t):
                put(("progress", "email", f"Fetching Steam… {i}/{t}"))
            rows = enrich_with_steam_reviews(products, progress_callback=progress)
            rows = apply_deal_filters(
                rows,
                score_type=params.get("score_type") or "All",
                score_value=params.get("score_value") or "",
                label_value=params.get("label_value") or "",
                min_reviews=params.get("min_reviews") or "",
                discount_value=params.get("discount_value") or "",
                price_value=params.get("price_value") or "",
                currency=currency,
                sale_end_type=params.get("sale_end_type") or "All",
                sale_end_value=params.get("sale_end_value") or "",
            )
            rows = _apply_publisher_filter(rows, params.get("publisher") or "")
            rows = _apply_developer_filter(rows, params.get("developer") or "")
            rows = _apply_tags_filter(rows, params.get("tags") or "")
            full_rows_for_cache = rows
            pool = rows
        put(("progress", "email", "Enriching…"))
        for p in pool:
            end_ms = _sale_end_ms(p)
            end_formatted = _format_offer_ends_est(end_ms)
            p["sale_end_display"] = ("Offer ends " + end_formatted) if end_formatted else ""
        block_games = _email_block_games(blocks, pool, index, currency=currency)
        featured_games = []
        for i, block in enumerate(blocks):
            if (block.get("type") or "").strip().lower() == "featured" and block_games and i < len(block_games) and block_games[i]:
                featured_games.extend(block_games[i])
        for p in featured_games:
            if p.get("steam_app_id") is not None and not (p.get("short_description") or "").strip():
                details = fetch_app_details_full(p["steam_app_id"], use_cache=True)
                if details and details.get("short_description"):
                    p["short_description"] = (details.get("short_description") or "").strip()
        put(("progress", "email", "Building HTML…"))
        currency = (params.get("currency") or "USD").strip() or "USD"
        try:
            coupon = max(0, min(50, float((params.get("coupon") or "0").strip() or 0)))
        except ValueError:
            coupon = 0
        show_val = (params.get("show_val") or "price").strip().lower() or "price"
        if show_val not in ("price", "discount", "both"):
            show_val = "price"
        options = {
            "currency": currency,
            "show_price": show_val in ("price", "both"),
            "show_both": show_val == "both",
            "coupon_percent": coupon,
        }
        def get_screenshots(app_id):
            out = fetch_app_details_full(app_id, use_cache=True)
            return (out or {}).get("screenshots") or []
        html = build_email_html(blocks, pool, options, get_screenshots=get_screenshots, block_games=block_games)
        status_msg = f"Preview built: {len(pool)} games, {len(blocks)} blocks."
        put(("done", "email_build", (pool, html, status_msg, full_rows_for_cache)))
    except Exception as e:
        put(("error", "email_build", e))


def _format_offer_ends_est(ms: int | None) -> str:
    """Format Unix ms (UTC) as EST date and time, e.g. 'Feb 20, 2026 11:59 PM EST'."""
    if ms is None:
        return ""
    try:
        est = ZoneInfo("America/New_York")
        dt = datetime.fromtimestamp(ms / 1000.0, tz=timezone.utc).astimezone(est)
        return dt.strftime("%b %d, %Y %I:%M %p EST")
    except (ValueError, OSError):
        return ""


def _lerp_hex(hex_a: str, hex_b: str, t: float) -> str:
    """Linear interpolate between two hex colors; t in [0, 1]."""
    def parse(h):
        h = h.lstrip("#")
        return tuple(int(h[i : i + 2], 16) for i in (0, 2, 4))
    a, b = parse(hex_a), parse(hex_b)
    r = int(a[0] + (b[0] - a[0]) * t)
    g = int(a[1] + (b[1] - a[1]) * t)
    bl = int(a[2] + (b[2] - a[2]) * t)
    return f"#{r:02x}{g:02x}{bl:02x}"


def parse_pasted_urls(text: str) -> list[str]:
    """Split pasted text into URLs (newline or comma separated), strip whitespace."""
    urls = []
    for part in text.replace(",", "\n").splitlines():
        u = part.strip()
        if u:
            urls.append(u)
    return urls


def _release_date_ms(row: dict) -> int | None:
    """Parse steam_release_date (e.g. 'Aug 21, 2012') to start-of-day UTC ms, or None."""
    s = (row.get("steam_release_date") or "").strip()
    if not s or s == "—":
        return None
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y"):
        try:
            dt = datetime.strptime(s, fmt)
            ts = datetime(dt.year, dt.month, dt.day, 0, 0, 0, 0, tzinfo=timezone.utc).timestamp()
            return int(ts * 1000)
        except (ValueError, TypeError):
            continue
    return None


def _apply_game_search_filter(rows: list[dict], query: str) -> list[dict]:
    """Filter rows by game title containing query (case-insensitive). Empty query = no filter."""
    q = (query or "").strip().lower()
    if not q:
        return rows
    return [r for r in rows if q in ((r.get("title") or "").strip().lower())]


def _apply_release_date_filter(
    rows: list[dict],
    filter_type: str,
    value_str: str,
) -> list[dict]:
    """Filter/sort by release date. filter_type: All, Newest, Oldest, By date."""
    if not filter_type or filter_type == "All":
        return rows
    if filter_type == "Newest":
        return sorted(rows, key=lambda r: (_release_date_ms(r) is None, -(_release_date_ms(r) or 0)))
    if filter_type == "Oldest":
        return sorted(rows, key=lambda r: (_release_date_ms(r) is None, _release_date_ms(r) or 0))
    if filter_type == "By date":
        mode, a, b = parse_sale_end_value(value_str)
        if mode == "":
            return rows
        if mode == "op":
            op_match = re.match(r"^(>=?|<=?|==?|!=)\s*(\d{4}-\d{2}-\d{2})$", (value_str or "").strip())
            if not op_match:
                return rows
            op, date_str = op_match.group(1), op_match.group(2)
            start_ms = date_str_to_start_of_day_ms(date_str)
            if start_ms is None:
                return rows
            end_ms = start_ms + ONE_DAY_MS - 1
            next_day_ms = start_ms + ONE_DAY_MS

            def ok(row):
                ms = _release_date_ms(row)
                if ms is None:
                    return False
                if op == "<": return ms < start_ms
                if op == "<=": return ms <= end_ms
                if op == ">": return ms >= next_day_ms
                if op == ">=": return ms >= start_ms
                if op == "==": return start_ms <= ms <= end_ms
                if op == "!=": return not (start_ms <= ms <= end_ms)
                return False
            return [r for r in rows if ok(r)]
        if mode == "range":
            def ok(row):
                ms = _release_date_ms(row)
                if ms is None:
                    return False
                return a <= ms <= b
            return [r for r in rows if ok(r)]
    return rows


def _apply_publisher_filter(rows: list[dict], query: str) -> list[dict]:
    """Filter rows by publisher containing query (case-insensitive). Empty query = no filter."""
    q = (query or "").strip().lower()
    if not q:
        return rows
    return [r for r in rows if q in ((r.get("steam_publisher") or "").strip().lower())]


def _apply_developer_filter(rows: list[dict], query: str) -> list[dict]:
    """Filter rows by developer containing query (case-insensitive). Empty query = no filter."""
    q = (query or "").strip().lower()
    if not q:
        return rows
    return [r for r in rows if q in ((r.get("steam_developer") or "").strip().lower())]


def _apply_tags_filter(rows: list[dict], query: str) -> list[dict]:
    """Filter rows by any tag containing query (case-insensitive). Empty query = no filter."""
    q = (query or "").strip().lower()
    if not q:
        return rows
    def matches(row):
        tags = row.get("steam_tags")
        if isinstance(tags, list):
            return any(q in (str(t).strip().lower()) for t in tags if t)
        s = (tags or "").strip().lower()
        return q in s
    return [r for r in rows if matches(r)]


def _game_used_key(g: dict) -> str | tuple[str, int] | None:
    """Return a hashable key for deduplication: normalized link or ('s', steam_app_id)."""
    link = (g.get("link") or "").strip()
    if link:
        return normalize_url(link)
    app_id = g.get("steam_app_id")
    if app_id is not None:
        return ("s", app_id)
    return None


def _game_pick_score(g: dict) -> float:
    """Score for stratified weighted sampling: rating, reviews, discount, owners, ccu. Returns 0..1-ish."""
    pct = g.get("steam_percent_positive")
    rating = (float(pct) / 100.0) if pct is not None else 0.0
    rev = g.get("steam_total_reviews") or 0
    reviews = min(1.0, math.log10(1 + rev) / 6.0) if rev else 0.0
    discount_pct = _discount_pct(g)
    discount = min(1.0, (float(discount_pct or 0) / 100.0))
    owners = g.get("steamspy_owners_estimate") or 0
    owners_n = min(1.0, math.log10(1 + owners) / 8.0) if owners else 0.0
    ccu = g.get("steamspy_ccu") or 0
    ccu_n = min(1.0, math.log10(1 + ccu) / 5.0) if ccu else 0.0
    return 0.25 * rating + 0.2 * reviews + 0.2 * discount + 0.2 * owners_n + 0.15 * ccu_n


def _weighted_sample(games: list[dict], n: int, score_fn: callable) -> list[dict]:
    """Sample n games without replacement; weights = score_fn(g). Criteria already applied to games."""
    if n <= 0 or not games:
        return []
    if n >= len(games):
        return list(games)
    work = [(g, max(1e-6, score_fn(g))) for g in games]
    out = []
    for _ in range(n):
        if not work:
            break
        total = sum(w for _, w in work)
        if total <= 0:
            break
        r = random.random() * total
        for idx, (game, w) in enumerate(work):
            r -= w
            if r <= 0:
                out.append(game)
                work.pop(idx)
                break
        else:
            out.append(work.pop(-1)[0])
    return out


def _email_block_games(blocks: list[dict], pool: list[dict], index: dict | None = None, currency: str = "USD") -> list[list[dict] | None]:
    """Build per-block game lists from pool. No game appears twice in the email: used set tracks assignments across blocks."""
    link_to_game = {normalize_url(g.get("link") or ""): g for g in pool if (g.get("link") or "").strip()}
    result: list[list[dict] | None] = [None] * len(blocks)
    used: set = set()
    for i, block in enumerate(blocks):
        btype = (block.get("type") or "").strip().lower()
        if btype not in ("deal_list", "featured"):
            continue
        cfg = block.get("config") or {}
        if btype == "deal_list":
            override_urls = cfg.get("override_urls")
            if override_urls and index is not None:
                products, _ = resolve_urls_to_products(index, override_urls)
                result[i] = [link_to_game[normalize_url(p.get("link") or "")] for p in products if normalize_url(p.get("link") or "") in link_to_game]
            else:
                override_ids = cfg.get("override_steam_ids")
                if override_ids:
                    id_to_game = {g.get("steam_app_id"): g for g in pool if g.get("steam_app_id") is not None}
                    result[i] = [id_to_game[aid] for aid in override_ids if aid in id_to_game]
                else:
                    filtered = list(pool)
                    filtered = _apply_publisher_filter(filtered, (cfg.get("publisher") or "").strip())
                    filtered = _apply_developer_filter(filtered, (cfg.get("developer") or "").strip())
                    filtered = _apply_tags_filter(filtered, (cfg.get("tags") or "").strip())
                    filtered = apply_price_filter(filtered, (cfg.get("price_value") or "").strip(), currency)
                    filtered = apply_discount_filter(filtered, (cfg.get("discount_value") or "").strip())
                    filtered = [g for g in filtered if _game_used_key(g) not in used]
                    n = max(0, int((cfg.get("games_count") or 4)))
                    result[i] = _weighted_sample(filtered, n, _game_pick_score)
            for g in result[i] or []:
                k = _game_used_key(g)
                if k is not None:
                    used.add(k)
        else:  # featured
            override_url = (cfg.get("override_url") or "").strip()
            if override_url and index is not None:
                products, _ = resolve_urls_to_products(index, [override_url])
                if products and normalize_url(products[0].get("link") or "") in link_to_game:
                    result[i] = [link_to_game[normalize_url(products[0].get("link") or "")]]
                else:
                    result[i] = []
            else:
                override_id = cfg.get("override_steam_id")
                if override_id is not None:
                    found = [g for g in pool if g.get("steam_app_id") == override_id]
                    result[i] = found[:1] if found else []
                else:
                    filtered = list(pool)
                    filtered = _apply_publisher_filter(filtered, (cfg.get("publisher") or "").strip())
                    filtered = _apply_developer_filter(filtered, (cfg.get("developer") or "").strip())
                    filtered = _apply_tags_filter(filtered, (cfg.get("tags") or "").strip())
                    filtered = apply_price_filter(filtered, (cfg.get("price_value") or "").strip(), currency)
                    filtered = apply_discount_filter(filtered, (cfg.get("discount_value") or "").strip())
                    filtered = [g for g in filtered if _game_used_key(g) not in used]
                    result[i] = _weighted_sample(filtered, 1, _game_pick_score)
            for g in result[i] or []:
                k = _game_used_key(g)
                if k is not None:
                    used.add(k)
    return result


class Application:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Reddit Deal Table Tool")
        self.root.minsize(500, 450)
        self._feed_items = None
        self._index = None
        self._on_sale_rows = []  # Enriched on-sale list for tab 2
        self._tab2_displayed_rows = []  # Last rows shown (for Copy URLs)
        self._worker_queue = queue.Queue()
        self._worker_busy = False
        self._build_ui()

    def _build_ui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Tab 1: Deal Table ---
        tab1 = ttk.Frame(notebook, padding=10)
        notebook.add(tab1, text="Deal Table")

        ttk.Label(tab1, text="Product URLs (one per line or comma-separated):").pack(anchor=tk.W)
        self.input_text = scrolledtext.ScrolledText(tab1, height=6, width=70, wrap=tk.WORD)
        self.input_text.pack(fill=tk.X, pady=(0, 8))

        btn_frame = ttk.Frame(tab1)
        btn_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Button(btn_frame, text="Load feed", command=self._load_feed).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frame, text="Build table", command=self._build_table).pack(side=tk.LEFT)

        ttk.Label(tab1, text="Currencies to show:").pack(anchor=tk.W, pady=(8, 4))
        curr_frame = ttk.Frame(tab1)
        curr_frame.pack(fill=tk.X)
        self.currency_vars = {}
        for i, code in enumerate(ALL_CURRENCIES):
            var = tk.BooleanVar(value=code in DEFAULT_CURRENCIES)
            self.currency_vars[code] = var
            cb = ttk.Checkbutton(curr_frame, text=CURRENCY_LABELS.get(code, code), variable=var)
            cb.grid(row=i // 6, column=i % 6, sticky=tk.W, padx=(0, 12), pady=2)
        ttk.Frame(tab1, height=8).pack()

        ttk.Label(tab1, text="Reddit markdown (copy and paste to Reddit):").pack(anchor=tk.W, pady=(8, 4))
        self.output_text = scrolledtext.ScrolledText(tab1, height=14, width=70, wrap=tk.NONE, state=tk.DISABLED)
        self.output_text.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        ttk.Button(tab1, text="Copy to clipboard", command=self._copy_output).pack(anchor=tk.W)

        # --- Tab 2: Deal Finder ---
        tab2 = ttk.Frame(notebook, padding=10)
        notebook.add(tab2, text="Deal Finder")

        ttk.Label(tab2, text="Load the feed, then fetch on-sale products and their Steam ratings.").pack(anchor=tk.W)
        btn2 = ttk.Frame(tab2)
        btn2.pack(fill=tk.X, pady=8)
        ttk.Button(btn2, text="Load feed", command=self._load_feed_tab2).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn2, text="Fetch on-sale list & Steam data", command=self._fetch_on_sale).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn2, text="Clear Steam cache", command=self._clear_steam_cache).pack(side=tk.LEFT)
        self.tab2_status = ttk.Label(tab2, text="")
        self.tab2_status.pack(anchor=tk.W, pady=(0, 8))

        # Search
        search_frame = ttk.Frame(tab2)
        search_frame.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(search_frame, text="Search game:").pack(side=tk.LEFT, padx=(0, 6))
        self.tab2_search_var = tk.StringVar(value="")
        search_entry = ttk.Entry(search_frame, textvariable=self.tab2_search_var, width=22)
        search_entry.pack(side=tk.LEFT, padx=(0, 8))
        search_entry.bind("<Return>", lambda e: self._apply_filters_tab2())

        # Filters (collapsible)
        self._filters_visible = True
        filters_outer = ttk.Frame(tab2)
        filters_outer.pack(fill=tk.X, pady=(8, 0))
        filters_header = ttk.Frame(filters_outer)
        filters_header.pack(fill=tk.X)
        ttk.Label(filters_header, text="Filters:").pack(side=tk.LEFT, padx=(0, 8))
        self.tab2_filters_toggle_btn = ttk.Button(filters_header, text="Hide filters \u25BC", width=14, command=self._toggle_filters_tab2)
        self.tab2_filters_toggle_btn.pack(side=tk.LEFT)
        filt_frame = ttk.Frame(filters_outer)
        filt_frame.pack(fill=tk.X, pady=(4, 0))
        self.tab2_filt_frame = filt_frame
        ttk.Label(filt_frame, text="Score:").grid(row=0, column=0, sticky=tk.W, padx=(0, 4))
        self.score_filter_type = tk.StringVar(value="All")
        score_combo = ttk.Combobox(filt_frame, textvariable=self.score_filter_type, width=12, state="readonly")
        score_combo["values"] = ("All", "Exact %", "Operator", "Label")
        score_combo.grid(row=0, column=1, sticky=tk.W, padx=(0, 8))
        self.score_filter_value = ttk.Entry(filt_frame, width=10)
        self.score_filter_value.grid(row=0, column=2, sticky=tk.W, padx=(0, 8))
        ttk.Label(filt_frame, text="(e.g. 75 or >75)").grid(row=0, column=3, sticky=tk.W, padx=(0, 12))
        self.label_filter = tk.StringVar(value=STEAM_LABEL_ORDER[0] if STEAM_LABEL_ORDER else "")
        label_combo = ttk.Combobox(filt_frame, textvariable=self.label_filter, width=22, state="readonly")
        label_combo["values"] = tuple(STEAM_LABEL_ORDER)
        label_combo.grid(row=0, column=4, sticky=tk.W, padx=(0, 8))
        ttk.Label(filt_frame, text="Min reviews:").grid(row=0, column=5, sticky=tk.W, padx=(16, 4))
        self.min_reviews_var = tk.StringVar(value="")
        ttk.Entry(filt_frame, textvariable=self.min_reviews_var, width=8).grid(row=0, column=6, sticky=tk.W, padx=(0, 8))
        ttk.Label(filt_frame, text="% off:").grid(row=0, column=7, sticky=tk.W, padx=(16, 4))
        self.discount_filter_var = tk.StringVar(value="")
        ttk.Entry(filt_frame, textvariable=self.discount_filter_var, width=10).grid(row=0, column=8, sticky=tk.W, padx=(0, 4))
        ttk.Label(filt_frame, text="(e.g. >50)").grid(row=0, column=9, sticky=tk.W, padx=(0, 8))
        ttk.Button(filt_frame, text="Apply filters", command=self._apply_filters_tab2).grid(row=0, column=10, padx=(8, 0))

        # Sale end filter (row 1)
        ttk.Label(filt_frame, text="Sale end:").grid(row=1, column=0, sticky=tk.W, padx=(0, 4), pady=(8, 0))
        self.sale_end_filter_type = tk.StringVar(value="All")
        sale_end_combo = ttk.Combobox(filt_frame, textvariable=self.sale_end_filter_type, width=14, state="readonly")
        sale_end_combo["values"] = ("All", "Ending Soon", "Ending Latest", "By date")
        sale_end_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        self.sale_end_filter_value = tk.StringVar(value="")
        ttk.Entry(filt_frame, textvariable=self.sale_end_filter_value, width=28).grid(row=1, column=2, columnspan=2, sticky=tk.W, padx=(0, 4), pady=(8, 0))
        ttk.Label(filt_frame, text="(e.g. <2026-03-01 or 2026-02-01..2026-02-28)").grid(row=1, column=4, columnspan=6, sticky=tk.W, padx=(0, 8), pady=(8, 0))

        # Release date filter (row 2)
        ttk.Label(filt_frame, text="Release date:").grid(row=2, column=0, sticky=tk.W, padx=(0, 4), pady=(8, 0))
        self.release_date_filter_type = tk.StringVar(value="All")
        release_date_combo = ttk.Combobox(filt_frame, textvariable=self.release_date_filter_type, width=14, state="readonly")
        release_date_combo["values"] = ("All", "Newest", "Oldest", "By date")
        release_date_combo.grid(row=2, column=1, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        self.release_date_filter_value = tk.StringVar(value="")
        ttk.Entry(filt_frame, textvariable=self.release_date_filter_value, width=28).grid(row=2, column=2, columnspan=2, sticky=tk.W, padx=(0, 4), pady=(8, 0))
        ttk.Label(filt_frame, text="(e.g. >=2020-01-01 or 2018-01-01..2022-12-31)").grid(row=2, column=4, columnspan=6, sticky=tk.W, padx=(0, 8), pady=(8, 0))

        # Publisher, Developer, Tags (row 3)
        ttk.Label(filt_frame, text="Publisher:").grid(row=3, column=0, sticky=tk.W, padx=(0, 4), pady=(8, 0))
        self.tab2_publisher_filter_var = tk.StringVar(value="")
        ttk.Entry(filt_frame, textvariable=self.tab2_publisher_filter_var, width=18).grid(row=3, column=1, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        ttk.Label(filt_frame, text="Developer:").grid(row=3, column=2, sticky=tk.W, padx=(16, 4), pady=(8, 0))
        self.tab2_developer_filter_var = tk.StringVar(value="")
        ttk.Entry(filt_frame, textvariable=self.tab2_developer_filter_var, width=18).grid(row=3, column=3, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        tags_frame = ttk.Frame(filt_frame)
        tags_frame.grid(row=3, column=4, columnspan=3, sticky=tk.W, padx=(16, 0), pady=(8, 0))
        ttk.Label(tags_frame, text="Tags:").pack(side=tk.LEFT, padx=(0, 4))
        self.tab2_tags_filter_var = tk.StringVar(value="")
        ttk.Entry(tags_frame, textvariable=self.tab2_tags_filter_var, width=22).pack(side=tk.LEFT)

        ttk.Label(tab2, text="Results (best rating first):").pack(anchor=tk.W, pady=(8, 4))
        self.tab2_sheet_frame = ttk.Frame(tab2)
        self.tab2_sheet_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        self.tab2_sheet_frame.grid_columnconfigure(0, weight=1)
        self.tab2_sheet_frame.grid_rowconfigure(0, weight=1)
        # Column width ratios (Game, Rating, Reviews, % Off, Release date, Sale end, Developer, Publisher, Tags) - sum 1.0
        self._tab2_column_ratios = [0.22, 0.14, 0.08, 0.06, 0.10, 0.10, 0.12, 0.12, 0.08]
        self.tab2_sheet = Sheet(
            self.tab2_sheet_frame,
            headers=["Game", "Rating", "Reviews", "% Off", "Release date", "Sale end", "Developer", "Publisher", "Tags"],
            show_row_index=False,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            height=400,
            default_column_width=80,
            table_wrap="w",
            alternate_color="#F0F4F8",
        )
        # Center text in all columns except Game (column 0)
        self.tab2_sheet.align_columns([1, 2, 3, 4, 5, 6, 7, 8], align="center")
        self.tab2_sheet.enable_bindings()
        self.tab2_sheet.bind("<Double-1>", self._on_tab2_sheet_double_click)
        self.tab2_sheet.grid(row=0, column=0, sticky="nswe")
        self.tab2_sheet_frame.bind("<Configure>", self._resize_tab2_columns)
        self.root.after(100, self._resize_tab2_columns)
        ttk.Label(tab2, text="Double-click a row to open the product page.", font=("TkDefaultFont", 8)).pack(anchor=tk.W)
        btn_frame = ttk.Frame(tab2)
        btn_frame.pack(anchor=tk.W)
        ttk.Button(btn_frame, text="Copy product URLs to clipboard", command=self._copy_tab2_urls).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frame, text="Export to Excel…", command=self._export_tab2_to_xlsx).pack(side=tk.LEFT)

        # --- Tab 3: Email Builder ---
        self._email_blocks = []
        self._email_game_pool = []
        tab3 = ttk.Frame(notebook, padding=10)
        notebook.add(tab3, text="Email Builder")

        # Left/top: settings and block list
        top_frame = ttk.Frame(tab3)
        top_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(top_frame, text="Game source:").grid(row=0, column=0, sticky=tk.W, padx=(0, 8))
        self.email_source_var = tk.StringVar(value="auto")
        ttk.Radiobutton(top_frame, text="Auto-pick by criteria", variable=self.email_source_var, value="auto").grid(row=0, column=1, sticky=tk.W, padx=(0, 16))
        ttk.Radiobutton(top_frame, text="Use my list (URLs below)", variable=self.email_source_var, value="list").grid(row=0, column=2, sticky=tk.W)
        ttk.Label(tab3, text="Product URLs (when using list):").pack(anchor=tk.W, pady=(4, 0))
        self.email_urls_text = scrolledtext.ScrolledText(tab3, height=3, width=70, wrap=tk.WORD)
        self.email_urls_text.pack(fill=tk.X, pady=(0, 8))

        # Criteria (for auto-pick) - compact row
        ttk.Label(tab3, text="Criteria (auto-pick): Score, min reviews, % off, sale end").pack(anchor=tk.W, pady=(8, 4))
        crit_frame = ttk.Frame(tab3)
        crit_frame.pack(fill=tk.X)
        ttk.Label(crit_frame, text="Score:").grid(row=0, column=0, sticky=tk.W, padx=(0, 4))
        self.email_score_type = tk.StringVar(value="All")
        email_score_combo = ttk.Combobox(crit_frame, textvariable=self.email_score_type, width=10, state="readonly")
        email_score_combo["values"] = ("All", "Exact %", "Operator", "Label")
        email_score_combo.grid(row=0, column=1, sticky=tk.W, padx=(0, 4))
        self.email_score_value = ttk.Entry(crit_frame, width=8)
        self.email_score_value.grid(row=0, column=2, sticky=tk.W, padx=(0, 8))
        self.email_label_value = tk.StringVar(value=STEAM_LABEL_ORDER[0] if STEAM_LABEL_ORDER else "")
        email_label_combo = ttk.Combobox(crit_frame, textvariable=self.email_label_value, width=18, state="readonly")
        email_label_combo["values"] = tuple(STEAM_LABEL_ORDER)
        email_label_combo.grid(row=0, column=3, sticky=tk.W, padx=(0, 8))
        ttk.Label(crit_frame, text="Min rev:").grid(row=0, column=4, sticky=tk.W, padx=(8, 4))
        self.email_min_reviews = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_min_reviews, width=6).grid(row=0, column=5, sticky=tk.W, padx=(0, 8))
        ttk.Label(crit_frame, text="% off:").grid(row=0, column=6, sticky=tk.W, padx=(8, 4))
        self.email_discount_value = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_discount_value, width=8).grid(row=0, column=7, sticky=tk.W, padx=(0, 4))
        ttk.Label(crit_frame, text="Sale end:").grid(row=0, column=8, sticky=tk.W, padx=(8, 4))
        self.email_sale_end_type = tk.StringVar(value="All")
        email_sale_end_combo = ttk.Combobox(crit_frame, textvariable=self.email_sale_end_type, width=12, state="readonly")
        email_sale_end_combo["values"] = ("All", "Ending Soon", "Ending Latest", "By date")
        email_sale_end_combo.grid(row=0, column=9, sticky=tk.W, padx=(0, 4))
        self.email_sale_end_value = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_sale_end_value, width=14).grid(row=0, column=10, sticky=tk.W)
        ttk.Label(crit_frame, text="Publisher:").grid(row=1, column=0, sticky=tk.W, padx=(0, 4), pady=(8, 0))
        self.email_publisher_var = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_publisher_var, width=14).grid(row=1, column=1, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        ttk.Label(crit_frame, text="Developer:").grid(row=1, column=2, sticky=tk.W, padx=(8, 4), pady=(8, 0))
        self.email_developer_var = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_developer_var, width=14).grid(row=1, column=3, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        ttk.Label(crit_frame, text="Tags:").grid(row=1, column=4, sticky=tk.W, padx=(8, 4), pady=(8, 0))
        self.email_tags_var = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_tags_var, width=18).grid(row=1, column=5, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        ttk.Label(crit_frame, text="Price (e.g. <6):").grid(row=1, column=6, sticky=tk.W, padx=(8, 4), pady=(8, 0))
        self.email_price_value = tk.StringVar(value="")
        ttk.Entry(crit_frame, textvariable=self.email_price_value, width=10).grid(row=1, column=7, sticky=tk.W, padx=(0, 8), pady=(8, 0))

        # Display options
        ttk.Label(tab3, text="Display:").pack(anchor=tk.W, pady=(8, 4))
        disp_frame = ttk.Frame(tab3)
        disp_frame.pack(fill=tk.X)
        ttk.Label(disp_frame, text="Show:").grid(row=0, column=0, sticky=tk.W, padx=(0, 4))
        self.email_show_var = tk.StringVar(value="price")
        ttk.Radiobutton(disp_frame, text="Price", variable=self.email_show_var, value="price").grid(row=0, column=1, sticky=tk.W, padx=(0, 12))
        ttk.Radiobutton(disp_frame, text="Discount %", variable=self.email_show_var, value="discount").grid(row=0, column=2, sticky=tk.W, padx=(0, 12))
        ttk.Radiobutton(disp_frame, text="Both", variable=self.email_show_var, value="both").grid(row=0, column=3, sticky=tk.W, padx=(0, 16))
        ttk.Label(disp_frame, text="Currency:").grid(row=0, column=4, sticky=tk.W, padx=(0, 4))
        self.email_currency_var = tk.StringVar(value="USD")
        curr_combo = ttk.Combobox(disp_frame, textvariable=self.email_currency_var, width=10, state="readonly")
        curr_combo["values"] = tuple(ALL_CURRENCIES)
        curr_combo.grid(row=0, column=5, sticky=tk.W, padx=(0, 16))
        ttk.Label(disp_frame, text="Coupon % off:").grid(row=0, column=6, sticky=tk.W, padx=(0, 4))
        self.email_coupon_var = tk.StringVar(value="0")
        ttk.Spinbox(disp_frame, from_=0, to=50, width=5, textvariable=self.email_coupon_var).grid(row=0, column=7, sticky=tk.W)

        # Block list
        ttk.Label(tab3, text="Blocks (order):").pack(anchor=tk.W, pady=(8, 4))
        block_btn_frame = ttk.Frame(tab3)
        block_btn_frame.pack(fill=tk.X)
        ttk.Button(block_btn_frame, text="Add block", command=self._email_add_block).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(block_btn_frame, text="Remove", command=self._email_remove_block).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(block_btn_frame, text="Move up", command=self._email_move_block_up).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(block_btn_frame, text="Move down", command=self._email_move_block_down).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(block_btn_frame, text="Edit", command=self._email_edit_block).pack(side=tk.LEFT)
        block_list_frame = ttk.Frame(tab3)
        block_list_frame.pack(fill=tk.X, pady=(0, 8))
        self.email_block_listbox = tk.Listbox(block_list_frame, height=8, width=50)
        self.email_block_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll = ttk.Scrollbar(block_list_frame, orient=tk.VERTICAL, command=self.email_block_listbox.yview)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.email_block_listbox.config(yscrollcommand=scroll.set)
        self.email_block_listbox.bind("<Double-1>", lambda e: self._email_edit_block())

        # Templates
        ttk.Label(tab3, text="Templates:").pack(anchor=tk.W, pady=(8, 4))
        template_btn_frame = ttk.Frame(tab3)
        template_btn_frame.pack(fill=tk.X)
        ttk.Button(template_btn_frame, text="Save as template…", command=self._email_save_template).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(template_btn_frame, text="Load template", command=self._email_load_template).pack(side=tk.LEFT)

        # Preview and export
        ttk.Label(tab3, text="Preview & export:").pack(anchor=tk.W, pady=(8, 4))
        export_frame = ttk.Frame(tab3)
        export_frame.pack(fill=tk.X)
        ttk.Button(export_frame, text="Load feed & build preview", command=self._email_build_preview).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(export_frame, text="Update preview", command=self._email_update_preview).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(export_frame, text="Export HTML…", command=self._email_export_html).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(export_frame, text="Open preview in browser", command=self._email_open_preview_browser).pack(side=tk.LEFT)
        self.email_status_var = tk.StringVar(value="")
        ttk.Label(tab3, textvariable=self.email_status_var).pack(anchor=tk.W)
        self._email_last_html = ""

    def _email_block_label(self, block: dict) -> str:
        btype = (block.get("type") or "").strip()
        cfg = block.get("config") or {}
        if btype == "deal_list":
            return f"Deal list ({cfg.get('games_count', 4)} games)"
        if btype == "featured":
            return "Featured (1 game)"
        if btype == "game_screenshots":
            p = cfg.get("product")
            title = (p.get("title") or "").strip()[:20] if p else "—"
            return f"Game screenshots ({title}…)"
        if btype == "title":
            return f"Title: {(cfg.get('text') or '')[:30]}…" if (cfg.get("text") or "").strip() else "Title"
        if btype == "button":
            return f"Button: {(cfg.get('text') or 'View more')[:20]}"
        return btype.capitalize()

    def _email_refresh_listbox(self):
        self.email_block_listbox.delete(0, tk.END)
        for b in self._email_blocks:
            self.email_block_listbox.insert(tk.END, self._email_block_label(b))

    def _email_add_block(self):
        types = ["header", "title", "deal_list", "featured", "text", "picture", "button", "game_screenshots", "footer"]
        menu = tk.Menu(self.root, tearoff=0)
        for t in types:
            menu.add_command(label=t.replace("_", " ").title(), command=lambda bt=t: self._email_do_add_block(bt))
        try:
            menu.tk_popup(self.root.winfo_pointerx(), self.root.winfo_pointery())
        finally:
            menu.grab_release()

    def _email_do_add_block(self, btype: str):
        block = {"type": btype, "config": {}}
        if btype == "deal_list":
            block["config"] = {"games_count": 4, "image_source": "feed", "capsule_size": "header", "show_titles": True, "show_rating": False, "show_reviews": False, "rating_style": "percent", "publisher": "", "developer": "", "tags": "", "price_value": "", "discount_value": "", "override_urls": []}
        elif btype == "featured":
            block["config"] = {"image_source": "feed", "capsule_size": "header", "show_titles": True, "show_rating": False, "show_reviews": False, "rating_style": "percent", "publisher": "", "developer": "", "tags": "", "price_value": "", "discount_value": "", "override_url": ""}
        elif btype == "game_screenshots":
            block["config"] = {}
        elif btype == "button":
            block["config"] = {"text": "View more", "url": ""}
        self._email_blocks.append(block)
        self._email_refresh_listbox()

    def _email_remove_block(self):
        sel = self.email_block_listbox.curselection()
        if not sel:
            return
        idx = int(sel[0])
        self._email_blocks.pop(idx)
        self._email_refresh_listbox()

    def _email_move_block_up(self):
        sel = self.email_block_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        idx = int(sel[0])
        self._email_blocks[idx], self._email_blocks[idx - 1] = self._email_blocks[idx - 1], self._email_blocks[idx]
        self._email_refresh_listbox()
        self.email_block_listbox.selection_set(idx - 1)

    def _email_move_block_down(self):
        sel = self.email_block_listbox.curselection()
        if not sel or sel[0] >= len(self._email_blocks) - 1:
            return
        idx = int(sel[0])
        self._email_blocks[idx], self._email_blocks[idx + 1] = self._email_blocks[idx + 1], self._email_blocks[idx]
        self._email_refresh_listbox()
        self.email_block_listbox.selection_set(idx + 1)

    def _email_edit_block(self):
        sel = self.email_block_listbox.curselection()
        if not sel:
            return
        idx = int(sel[0])
        block = self._email_blocks[idx]
        btype = (block.get("type") or "").strip().lower()
        cfg = block.get("config") or {}
        win = tk.Toplevel(self.root)
        win.title(f"Edit {btype} block")
        win.transient(self.root)
        f = ttk.Frame(win, padding=10)
        f.pack(fill=tk.BOTH, expand=True)
        f.grid_columnconfigure(0, minsize=150)
        entries = {}

        if btype == "header":
            ttk.Label(f, text="Title:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["title"] = ttk.Entry(f, width=40)
            entries["title"].insert(0, cfg.get("title") or "")
            entries["title"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Logo URL:").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["logo_url"] = ttk.Entry(f, width=40)
            entries["logo_url"].insert(0, cfg.get("logo_url") or "")
            entries["logo_url"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Link:").grid(row=2, column=0, sticky=tk.W, pady=2)
            entries["link"] = ttk.Entry(f, width=40)
            entries["link"].insert(0, cfg.get("link") or "")
            entries["link"].grid(row=2, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="View in browser URL:").grid(row=3, column=0, sticky=tk.W, pady=2)
            entries["view_in_browser_url"] = ttk.Entry(f, width=40)
            entries["view_in_browser_url"].insert(0, cfg.get("view_in_browser_url") or "")
            entries["view_in_browser_url"].grid(row=3, column=1, sticky=tk.W, pady=2, padx=(4, 0))
        elif btype == "title":
            ttk.Label(f, text="Text:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["text"] = ttk.Entry(f, width=50)
            entries["text"].insert(0, cfg.get("text") or "")
            entries["text"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
        elif btype == "deal_list":
            ttk.Label(f, text="Games count:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["games_count"] = ttk.Entry(f, width=6)
            entries["games_count"].insert(0, str(cfg.get("games_count") or 4))
            entries["games_count"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Section title:").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["section_title"] = ttk.Entry(f, width=40)
            entries["section_title"].insert(0, cfg.get("section_title") or "")
            entries["section_title"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Image source:").grid(row=2, column=0, sticky=tk.W, pady=2)
            entries["image_source"] = tk.StringVar(value=cfg.get("image_source") or "feed")
            ttk.Radiobutton(f, text="Product feed cover", variable=entries["image_source"], value="feed").grid(row=2, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Radiobutton(f, text="Steam capsule", variable=entries["image_source"], value="steam_capsule").grid(row=3, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="Capsule size:").grid(row=4, column=0, sticky=tk.W, pady=2)
            entries["capsule_size"] = ttk.Combobox(f, width=12, state="readonly")
            entries["capsule_size"]["values"] = ("header", "capsule_sm", "capsule_md", "capsule_616x353")
            entries["capsule_size"].set(cfg.get("capsule_size") or "header")
            entries["capsule_size"].grid(row=4, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            entries["show_titles"] = tk.BooleanVar(value=cfg.get("show_titles", True))
            ttk.Checkbutton(f, text="Show game titles", variable=entries["show_titles"]).grid(row=5, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="Steam reviews:").grid(row=6, column=0, sticky=tk.W, pady=(8, 2))
            entries["show_rating"] = tk.BooleanVar(value=cfg.get("show_rating", False))
            entries["show_reviews"] = tk.BooleanVar(value=cfg.get("show_reviews", False))
            ttk.Checkbutton(f, text="Show rating", variable=entries["show_rating"]).grid(row=6, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Checkbutton(f, text="Show review count", variable=entries["show_reviews"]).grid(row=7, column=1, sticky=tk.W, padx=(4, 0))
            entries["rating_style"] = tk.StringVar(value=cfg.get("rating_style") or "percent")
            ttk.Radiobutton(f, text="Rating as: %", variable=entries["rating_style"], value="percent").grid(row=8, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Radiobutton(f, text="Rating as: label (e.g. Very Positive)", variable=entries["rating_style"], value="label").grid(row=9, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="Limit to (optional):").grid(row=10, column=0, sticky=tk.W, pady=(12, 2))
            ttk.Label(f, text="Publisher:").grid(row=11, column=0, sticky=tk.W, padx=(0, 4))
            entries["block_publisher"] = ttk.Entry(f, width=18)
            entries["block_publisher"].insert(0, (cfg.get("publisher") or "").strip())
            entries["block_publisher"].grid(row=11, column=1, sticky=tk.W, padx=(0, 4))
            ttk.Label(f, text="Developer:").grid(row=11, column=2, sticky=tk.W, padx=(4, 4))
            entries["block_developer"] = ttk.Entry(f, width=18)
            entries["block_developer"].insert(0, (cfg.get("developer") or "").strip())
            entries["block_developer"].grid(row=11, column=3, sticky=tk.W, padx=(0, 4))
            ttk.Label(f, text="Tags:").grid(row=11, column=4, sticky=tk.W, padx=(4, 4))
            entries["block_tags"] = ttk.Entry(f, width=16)
            entries["block_tags"].insert(0, (cfg.get("tags") or "").strip())
            entries["block_tags"].grid(row=11, column=5, sticky=tk.W, padx=(0, 4))
            ttk.Label(f, text="Price (e.g. <6):").grid(row=12, column=0, sticky=tk.W, pady=(8, 2))
            entries["block_price_value"] = ttk.Entry(f, width=10)
            entries["block_price_value"].insert(0, (cfg.get("price_value") or "").strip())
            entries["block_price_value"].grid(row=12, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="% off (e.g. >50):").grid(row=12, column=2, sticky=tk.W, padx=(8, 4), pady=(8, 2))
            entries["block_discount_value"] = ttk.Entry(f, width=10)
            entries["block_discount_value"].insert(0, (cfg.get("discount_value") or "").strip())
            entries["block_discount_value"].grid(row=12, column=3, sticky=tk.W, padx=(4, 0))
            url_lbl = ttk.Label(f, text="Product URLs (one per line or comma-separated; empty = auto):", wraplength=250)
            url_lbl.grid(row=13, column=0, sticky=tk.W, pady=(8, 2))
            entries["override_urls_text"] = scrolledtext.ScrolledText(f, width=50, height=5, wrap=tk.WORD)
            override_urls_dl = cfg.get("override_urls") or []
            entries["override_urls_text"].insert("1.0", "\n".join(str(u) for u in override_urls_dl if u))
            entries["override_urls_text"].grid(row=14, column=0, columnspan=6, sticky=tk.W, pady=(0, 4))
        elif btype == "featured":
            ttk.Label(f, text="Description:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["description"] = scrolledtext.ScrolledText(f, width=40, height=4, wrap=tk.WORD)
            entries["description"].insert("1.0", cfg.get("description") or "")
            entries["description"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Offer ends (e.g. Offer ends Nov 2):").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["offer_ends"] = ttk.Entry(f, width=40)
            entries["offer_ends"].insert(0, cfg.get("offer_ends") or "")
            entries["offer_ends"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Image source:").grid(row=2, column=0, sticky=tk.W, pady=2)
            entries["image_source"] = tk.StringVar(value=cfg.get("image_source") or "feed")
            ttk.Radiobutton(f, text="Product feed cover", variable=entries["image_source"], value="feed").grid(row=2, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Radiobutton(f, text="Steam capsule", variable=entries["image_source"], value="steam_capsule").grid(row=3, column=1, sticky=tk.W, padx=(4, 0))
            entries["capsule_size"] = ttk.Combobox(f, width=12, state="readonly")
            entries["capsule_size"]["values"] = ("header", "capsule_sm", "capsule_md", "capsule_616x353")
            entries["capsule_size"].set(cfg.get("capsule_size") or "header")
            entries["capsule_size"].grid(row=4, column=1, sticky=tk.W, padx=(4, 0))
            entries["show_titles"] = tk.BooleanVar(value=cfg.get("show_titles", True))
            ttk.Checkbutton(f, text="Show game titles", variable=entries["show_titles"]).grid(row=5, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="Steam reviews:").grid(row=6, column=0, sticky=tk.W, pady=(8, 2))
            entries["show_rating"] = tk.BooleanVar(value=cfg.get("show_rating", False))
            entries["show_reviews"] = tk.BooleanVar(value=cfg.get("show_reviews", False))
            ttk.Checkbutton(f, text="Show rating", variable=entries["show_rating"]).grid(row=6, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Checkbutton(f, text="Show review count", variable=entries["show_reviews"]).grid(row=7, column=1, sticky=tk.W, padx=(4, 0))
            entries["rating_style"] = tk.StringVar(value=cfg.get("rating_style") or "percent")
            ttk.Radiobutton(f, text="Rating as: %", variable=entries["rating_style"], value="percent").grid(row=8, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Radiobutton(f, text="Rating as: label (e.g. Very Positive)", variable=entries["rating_style"], value="label").grid(row=9, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="Limit to (optional):").grid(row=10, column=0, sticky=tk.W, pady=(12, 2))
            ttk.Label(f, text="Publisher:").grid(row=11, column=0, sticky=tk.W, padx=(0, 2))
            entries["block_publisher_f"] = ttk.Entry(f, width=18)
            entries["block_publisher_f"].insert(0, (cfg.get("publisher") or "").strip())
            entries["block_publisher_f"].grid(row=11, column=1, sticky=tk.W, padx=(0, 4))
            ttk.Label(f, text="Developer:").grid(row=11, column=2, sticky=tk.W, padx=(4, 4))
            entries["block_developer_f"] = ttk.Entry(f, width=18)
            entries["block_developer_f"].insert(0, (cfg.get("developer") or "").strip())
            entries["block_developer_f"].grid(row=11, column=3, sticky=tk.W, padx=(0, 4))
            ttk.Label(f, text="Tags:").grid(row=11, column=4, sticky=tk.W, padx=(4, 4))
            entries["block_tags_f"] = ttk.Entry(f, width=16)
            entries["block_tags_f"].insert(0, (cfg.get("tags") or "").strip())
            entries["block_tags_f"].grid(row=11, column=5, sticky=tk.W, padx=(0, 4))
            ttk.Label(f, text="Price (e.g. <6):").grid(row=12, column=0, sticky=tk.W, pady=(8, 2))
            entries["block_price_value_f"] = ttk.Entry(f, width=10)
            entries["block_price_value_f"].insert(0, (cfg.get("price_value") or "").strip())
            entries["block_price_value_f"].grid(row=12, column=1, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="% off (e.g. >50):").grid(row=12, column=2, sticky=tk.W, padx=(8, 4), pady=(8, 2))
            entries["block_discount_value_f"] = ttk.Entry(f, width=10)
            entries["block_discount_value_f"].insert(0, (cfg.get("discount_value") or "").strip())
            entries["block_discount_value_f"].grid(row=12, column=3, sticky=tk.W, padx=(4, 0))
            ttk.Label(f, text="Product URL (empty = auto):").grid(row=13, column=0, sticky=tk.W, pady=(8, 2))
            entries["override_url_text"] = ttk.Entry(f, width=50)
            entries["override_url_text"].insert(0, (cfg.get("override_url") or "").strip())
            entries["override_url_text"].grid(row=14, column=0, columnspan=6, sticky=tk.W, pady=(0, 4))
        elif btype == "text":
            ttk.Label(f, text="Content (HTML allowed):").grid(row=0, column=0, sticky=tk.NW, pady=2)
            entries["content"] = scrolledtext.ScrolledText(f, width=50, height=6, wrap=tk.WORD)
            entries["content"].insert("1.0", cfg.get("content") or "")
            entries["content"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
        elif btype == "picture":
            ttk.Label(f, text="Image URL:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["image_url"] = ttk.Entry(f, width=50)
            entries["image_url"].insert(0, cfg.get("image_url") or "")
            entries["image_url"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Link URL:").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["link_url"] = ttk.Entry(f, width=50)
            entries["link_url"].insert(0, cfg.get("link_url") or "")
            entries["link_url"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Alt text:").grid(row=2, column=0, sticky=tk.W, pady=2)
            entries["alt"] = ttk.Entry(f, width=30)
            entries["alt"].insert(0, cfg.get("alt") or "")
            entries["alt"].grid(row=2, column=1, sticky=tk.W, pady=2, padx=(4, 0))
        elif btype == "button":
            ttk.Label(f, text="Button text:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["text"] = ttk.Entry(f, width=30)
            entries["text"].insert(0, cfg.get("text") or "View more")
            entries["text"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="URL:").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["url"] = ttk.Entry(f, width=50)
            entries["url"].insert(0, cfg.get("url") or "")
            entries["url"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
        elif btype == "game_screenshots":
            ttk.Label(f, text="Game (from pool):").grid(row=0, column=0, sticky=tk.W, pady=2)
            pool = self._email_game_pool
            titles = [(p.get("title") or "").strip() or "—" for p in pool]
            entries["game_index"] = tk.StringVar(value=str(cfg.get("game_index", 0)) if pool else "0")
            game_combo = ttk.Combobox(f, textvariable=entries["game_index"], width=35, state="readonly")
            game_combo["values"] = [f"{i}: {t[:40]}" for i, t in enumerate(titles)] if titles else ["0: (no pool)"]
            game_combo.grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            if pool and 0 <= (cfg.get("game_index") or 0) < len(pool):
                game_combo.set(f"{cfg.get('game_index', 0)}: {(pool[cfg.get('game_index', 0)].get('title') or '')[:40]}")
            ttk.Label(f, text="Caption:").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["caption"] = ttk.Entry(f, width=40)
            entries["caption"].insert(0, cfg.get("caption") or "")
            entries["caption"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
        elif btype == "footer":
            ttk.Label(f, text="Unsubscribe URL:").grid(row=0, column=0, sticky=tk.W, pady=2)
            entries["unsubscribe_url"] = ttk.Entry(f, width=50)
            entries["unsubscribe_url"].insert(0, cfg.get("unsubscribe_url") or "")
            entries["unsubscribe_url"].grid(row=0, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Privacy URL:").grid(row=1, column=0, sticky=tk.W, pady=2)
            entries["privacy_url"] = ttk.Entry(f, width=50)
            entries["privacy_url"].insert(0, cfg.get("privacy_url") or "")
            entries["privacy_url"].grid(row=1, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Terms URL:").grid(row=2, column=0, sticky=tk.W, pady=2)
            entries["terms_url"] = ttk.Entry(f, width=50)
            entries["terms_url"].insert(0, cfg.get("terms_url") or "")
            entries["terms_url"].grid(row=2, column=1, sticky=tk.W, pady=2, padx=(4, 0))
            ttk.Label(f, text="Address:").grid(row=3, column=0, sticky=tk.W, pady=2)
            entries["address"] = ttk.Entry(f, width=50)
            entries["address"].insert(0, cfg.get("address") or "")
            entries["address"].grid(row=3, column=1, sticky=tk.W, pady=2, padx=(4, 0))

        def save():
            new_cfg = dict(cfg)
            if btype == "header":
                new_cfg["title"] = entries["title"].get().strip()
                new_cfg["logo_url"] = entries["logo_url"].get().strip()
                new_cfg["link"] = entries["link"].get().strip()
                new_cfg["view_in_browser_url"] = entries["view_in_browser_url"].get().strip()
            elif btype == "title":
                new_cfg["text"] = entries["text"].get().strip()
            elif btype == "deal_list":
                try:
                    new_cfg["games_count"] = int(entries["games_count"].get().strip()) or 4
                except ValueError:
                    new_cfg["games_count"] = 4
                new_cfg["section_title"] = entries["section_title"].get().strip()
                new_cfg["publisher"] = (entries.get("block_publisher") and entries["block_publisher"].get().strip()) or ""
                new_cfg["developer"] = (entries.get("block_developer") and entries["block_developer"].get().strip()) or ""
                new_cfg["tags"] = (entries.get("block_tags") and entries["block_tags"].get().strip()) or ""
                new_cfg["price_value"] = (entries.get("block_price_value") and entries["block_price_value"].get().strip()) or ""
                new_cfg["discount_value"] = (entries.get("block_discount_value") and entries["block_discount_value"].get().strip()) or ""
                urls_dl = parse_pasted_urls(entries.get("override_urls_text") and entries["override_urls_text"].get("1.0", tk.END) or "")
                new_cfg["override_urls"] = urls_dl
                new_cfg["image_source"] = entries["image_source"].get().strip() or "feed"
                new_cfg["capsule_size"] = entries["capsule_size"].get().strip() or "header"
                new_cfg["show_titles"] = entries["show_titles"].get()
                new_cfg["show_rating"] = entries["show_rating"].get()
                new_cfg["show_reviews"] = entries["show_reviews"].get()
                rs = (entries["rating_style"].get() or "percent").strip().lower()
                new_cfg["rating_style"] = rs if rs in ("percent", "label") else "percent"
            elif btype == "featured":
                new_cfg["description"] = entries["description"].get("1.0", tk.END).strip()
                new_cfg["offer_ends"] = entries["offer_ends"].get().strip()
                new_cfg["publisher"] = (entries.get("block_publisher_f") and entries["block_publisher_f"].get().strip()) or ""
                new_cfg["developer"] = (entries.get("block_developer_f") and entries["block_developer_f"].get().strip()) or ""
                new_cfg["tags"] = (entries.get("block_tags_f") and entries["block_tags_f"].get().strip()) or ""
                new_cfg["price_value"] = (entries.get("block_price_value_f") and entries["block_price_value_f"].get().strip()) or ""
                new_cfg["discount_value"] = (entries.get("block_discount_value_f") and entries["block_discount_value_f"].get().strip()) or ""
                new_cfg["override_url"] = (entries.get("override_url_text") and entries["override_url_text"].get().strip()) or ""
                new_cfg["image_source"] = entries["image_source"].get().strip() or "feed"
                new_cfg["capsule_size"] = entries["capsule_size"].get().strip() or "header"
                new_cfg["show_titles"] = entries["show_titles"].get()
                new_cfg["show_rating"] = entries["show_rating"].get()
                new_cfg["show_reviews"] = entries["show_reviews"].get()
                rs = (entries["rating_style"].get() or "percent").strip().lower()
                new_cfg["rating_style"] = rs if rs in ("percent", "label") else "percent"
            elif btype == "text":
                new_cfg["content"] = entries["content"].get("1.0", tk.END).strip()
            elif btype == "picture":
                new_cfg["image_url"] = entries["image_url"].get().strip()
                new_cfg["link_url"] = entries["link_url"].get().strip()
                new_cfg["alt"] = entries["alt"].get().strip()
            elif btype == "button":
                new_cfg["text"] = entries["text"].get().strip() or "View more"
                new_cfg["url"] = entries["url"].get().strip()
            elif btype == "game_screenshots":
                try:
                    idx_str = entries["game_index"].get().strip().split(":")[0]
                    new_cfg["game_index"] = int(idx_str)
                except (ValueError, IndexError):
                    new_cfg["game_index"] = 0
                if self._email_game_pool and 0 <= new_cfg["game_index"] < len(self._email_game_pool):
                    new_cfg["product"] = self._email_game_pool[new_cfg["game_index"]]
                new_cfg["caption"] = entries["caption"].get().strip()
            elif btype == "footer":
                new_cfg["unsubscribe_url"] = entries["unsubscribe_url"].get().strip()
                new_cfg["privacy_url"] = entries["privacy_url"].get().strip()
                new_cfg["terms_url"] = entries["terms_url"].get().strip()
                new_cfg["address"] = entries["address"].get().strip()
            block["config"] = new_cfg
            self._email_refresh_listbox()
            win.destroy()

        ttk.Button(win, text="Save", command=save).pack(pady=(10, 0))
        win.update_idletasks()
        w = win.winfo_reqwidth()
        h = win.winfo_reqheight()
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_w = self.root.winfo_width()
        root_h = self.root.winfo_height()
        x = root_x + max(0, (root_w - w) // 2)
        y = root_y + max(0, (root_h - h) // 2)
        win.geometry(f"+{x}+{y}")

    def _email_save_template(self):
        """Save current blocks and display options as a named template (Option A: one JSON file per template)."""
        name = simpledialog.askstring("Save template", "Template name:", parent=self.root)
        if not name or not name.strip():
            return
        # Sanitize to a safe filename (alnum + underscore)
        safe = "".join(c if c.isalnum() or c in " _-" else "" for c in name.strip()).replace(" ", "_").strip("_")
        if not safe:
            safe = "template"
        os.makedirs(EMAIL_TEMPLATES_DIR, exist_ok=True)
        # Build blocks for storage: strip runtime-only 'product' from game_screenshots
        blocks = []
        for b in self._email_blocks:
            blk = {"type": b.get("type") or "", "config": dict(b.get("config") or {})}
            if blk["type"] == "game_screenshots" and "product" in blk["config"]:
                del blk["config"]["product"]
            blocks.append(blk)
        show_val = (self.email_show_var.get() or "price").strip().lower() or "price"
        if show_val not in ("price", "discount", "both"):
            show_val = "price"
        try:
            coupon = max(0, min(50, float(self.email_coupon_var.get().strip() or 0)))
        except ValueError:
            coupon = 0
        display = {
            "currency": (self.email_currency_var.get() or "USD").strip() or "USD",
            "show": show_val,
            "coupon_percent": coupon,
        }
        path = os.path.join(EMAIL_TEMPLATES_DIR, safe + ".json")
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"blocks": blocks, "display": display}, f, indent=2)
            messagebox.showinfo("Template saved", f"Template saved as '{safe}'.")
        except OSError as e:
            messagebox.showerror("Error", f"Could not save template: {e}")

    def _email_load_template(self):
        """Load a saved template: replace blocks and display options (Option A: list JSON files in templates dir)."""
        if not os.path.isdir(EMAIL_TEMPLATES_DIR):
            messagebox.showinfo("Load template", "No templates saved yet. Use 'Save as template…' first.")
            return
        files = sorted(f for f in os.listdir(EMAIL_TEMPLATES_DIR) if f.endswith(".json"))
        if not files:
            messagebox.showinfo("Load template", "No saved templates.")
            return
        names = [os.path.splitext(f)[0] for f in files]
        win = tk.Toplevel(self.root)
        win.title("Load template")
        win.transient(self.root)
        frm = ttk.Frame(win, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text="Choose a template:").pack(anchor=tk.W)
        lb = tk.Listbox(frm, height=min(12, len(names)), width=40)
        lb.pack(fill=tk.BOTH, expand=True, pady=(4, 8))
        for n in names:
            lb.insert(tk.END, n)
        if names:
            lb.selection_set(0)
        btn_frm = ttk.Frame(frm)
        btn_frm.pack(fill=tk.X)

        def load():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("Load template", "Select a template.")
                return
            name = names[int(sel[0])]
            path = os.path.join(EMAIL_TEMPLATES_DIR, name + ".json")
            try:
                with open(path, encoding="utf-8") as fp:
                    data = json.load(fp)
            except (OSError, json.JSONDecodeError) as e:
                messagebox.showerror("Error", f"Could not load template: {e}")
                return
            blocks = data.get("blocks") or []
            if not isinstance(blocks, list):
                messagebox.showerror("Error", "Invalid template: 'blocks' must be a list.")
                return
            self._email_blocks = [{"type": b.get("type") or "", "config": dict(b.get("config") or {})} for b in blocks]
            disp = data.get("display") or {}
            if isinstance(disp, dict):
                if "currency" in disp:
                    self.email_currency_var.set(disp.get("currency") or "USD")
                if "show" in disp and (disp.get("show") or "").strip() in ("price", "discount", "both"):
                    self.email_show_var.set((disp.get("show") or "price").strip())
                if "coupon_percent" in disp:
                    try:
                        v = max(0, min(50, float(disp.get("coupon_percent", 0))))
                        self.email_coupon_var.set(str(int(v) if v == int(v) else v))
                    except (TypeError, ValueError):
                        pass
            self._email_refresh_listbox()
            win.destroy()

        ttk.Button(btn_frm, text="Load", command=load).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frm, text="Cancel", command=win.destroy).pack(side=tk.LEFT)

    def _email_get_game_pool(self) -> list[dict]:
        """Build game pool for email: from URL list or auto-pick with filters (full list, no cap)."""
        if self._index is None:
            self._load_feed()
        if self._index is None:
            return []
        if self.email_source_var.get() == "list":
            urls = parse_pasted_urls(self.email_urls_text.get("1.0", tk.END))
            products, _ = resolve_urls_to_products(self._index, urls)
            return list(products)
        products = get_on_sale_products(self._index, resolve_steam_by_name=True)
        rows = enrich_with_steam_reviews(products, progress_callback=None)
        currency = (self.email_currency_var.get() or "USD").strip() or "USD"
        rows = apply_deal_filters(
            rows,
            score_type=self.email_score_type.get(),
            score_value=self.email_score_value.get(),
            label_value=self.email_label_value.get(),
            min_reviews=self.email_min_reviews.get(),
            discount_value=self.email_discount_value.get(),
            price_value=self.email_price_value.get(),
            currency=currency,
            sale_end_type=self.email_sale_end_type.get(),
            sale_end_value=self.email_sale_end_value.get(),
        )
        rows = _apply_publisher_filter(rows, self.email_publisher_var.get() or "")
        rows = _apply_developer_filter(rows, self.email_developer_var.get() or "")
        rows = _apply_tags_filter(rows, self.email_tags_var.get() or "")
        return rows

    def _email_build_preview(self):
        if self._worker_busy:
            messagebox.showwarning("Please wait", "Another operation is in progress.")
            return
        if self._index is None:
            messagebox.showwarning("Load feed first", "Load the feed from the Deal Table or Deal Finder tab first.")
            return
        self.email_status_var.set("Building…")
        self._worker_busy = True
        self.root.after(50, self._process_worker_queue)
        index = self._index
        params = {
            "source": self.email_source_var.get(),
            "urls_text": self.email_urls_text.get("1.0", tk.END),
            "score_type": self.email_score_type.get(),
            "score_value": self.email_score_value.get(),
            "label_value": self.email_label_value.get(),
            "min_reviews": self.email_min_reviews.get(),
            "discount_value": self.email_discount_value.get(),
            "price_value": self.email_price_value.get(),
            "sale_end_type": self.email_sale_end_type.get(),
            "sale_end_value": self.email_sale_end_value.get(),
            "publisher": self.email_publisher_var.get(),
            "developer": self.email_developer_var.get(),
            "tags": self.email_tags_var.get(),
            "currency": self.email_currency_var.get(),
            "coupon": self.email_coupon_var.get(),
            "show_val": self.email_show_var.get(),
        }
        blocks = list(self._email_blocks)
        source = (self.email_source_var.get() or "auto").strip() or "auto"
        pre_enriched = self._on_sale_rows if (source == "auto" and self._on_sale_rows) else None

        def work():
            _email_build_worker(self._worker_queue, index, params, blocks, pre_enriched_rows=pre_enriched)

        threading.Thread(target=work, daemon=True).start()

    def _email_update_preview(self):
        """Rebuild email HTML from existing game pool and current blocks/options. No feed or Steam re-fetch."""
        if not self._email_game_pool:
            messagebox.showwarning("No preview", "Build preview first (Load feed & build preview), then you can use Update preview for quick changes.")
            return
        self.email_status_var.set("Updating…")
        self.root.update()
        try:
            pool = self._email_game_pool
            for p in pool:
                end_ms = _sale_end_ms(p)
                end_formatted = _format_offer_ends_est(end_ms)
                p["sale_end_display"] = ("Offer ends " + end_formatted) if end_formatted else ""
            currency = (self.email_currency_var.get() or "USD").strip() or "USD"
            try:
                coupon = max(0, min(50, float(self.email_coupon_var.get().strip() or 0)))
            except ValueError:
                coupon = 0
            show_val = (self.email_show_var.get() or "price").strip().lower() or "price"
            if show_val not in ("price", "discount", "both"):
                show_val = "price"
            options = {
                "currency": currency,
                "show_price": show_val in ("price", "both"),
                "show_both": show_val == "both",
                "coupon_percent": coupon,
            }
            def get_screenshots(app_id):
                out = fetch_app_details_full(app_id, use_cache=True)
                return (out or {}).get("screenshots") or []
            block_games = _email_block_games(self._email_blocks, pool, self._index, currency=currency)
            html = build_email_html(
                self._email_blocks,
                pool,
                options,
                get_screenshots=get_screenshots,
                block_games=block_games,
            )
            self._email_last_html = html
            self.email_status_var.set(f"Preview updated: {len(pool)} games, {len(self._email_blocks)} blocks (no feed refresh).")
        except Exception as e:
            self.email_status_var.set("")
            messagebox.showerror("Error", str(e))
            import traceback
            traceback.print_exc()

    def _email_export_html(self):
        if not self._email_last_html:
            messagebox.showwarning("No preview", "Build preview first (Load feed & build preview), then export.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".html",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self._email_last_html)
            messagebox.showinfo("Exported", f"Saved to {path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _email_open_preview_browser(self):
        if not self._email_last_html:
            messagebox.showwarning("No preview", "Build preview first (Load feed & build preview), then open in browser.")
            return
        import tempfile
        with tempfile.NamedTemporaryFile(mode="w", suffix=".html", delete=False, encoding="utf-8") as f:
            f.write(self._email_last_html)
            path = f.name
        if sys.platform == "win32":
            os.startfile(path)
        else:
            webbrowser.open("file://" + path.replace("\\", "/"))

    def _get_selected_currencies(self) -> list[str]:
        return [c for c in ALL_CURRENCIES if self.currency_vars[c].get()]

    def _process_worker_queue(self):
        """Process messages from worker thread (progress, done, error). Must run on main thread."""
        while True:
            try:
                msg = self._worker_queue.get_nowait()
            except queue.Empty:
                break
            kind = msg[0] if isinstance(msg, (list, tuple)) else msg
            if kind == "progress":
                _, target, text = msg
                if target == "tab2":
                    self.tab2_status.config(text=text)
                elif target == "email":
                    self.email_status_var.set(text)
            elif kind == "done":
                _, op, payload = msg
                if op == "load_feed":
                    self._feed_items, self._index = payload
                    self.input_text.config(cursor="")
                    messagebox.showinfo("Feed loaded", f"Loaded {len(self._feed_items)} variants ({len(self._index)} products).")
                elif op == "fetch_on_sale":
                    self._on_sale_rows = payload
                    self._apply_filters_tab2()
                    self.tab2_status.config(text=f"Done. {len(self._tab2_displayed_rows)} products (after filters).")
                elif op == "email_build":
                    pool, html, status_msg = payload[0], payload[1], payload[2]
                    self._email_game_pool = pool
                    self._email_last_html = html
                    self.email_status_var.set(status_msg)
                    if len(payload) >= 4 and payload[3] is not None:
                        self._on_sale_rows = payload[3]
                    if self._index is not None:
                        currency = (self.email_currency_var.get() or "USD").strip() or "USD"
                        block_games = _email_block_games(self._email_blocks, pool, self._index, currency=currency)
                        for i, block in enumerate(self._email_blocks):
                            btype = (block.get("type") or "").strip().lower()
                            if btype not in ("deal_list", "featured"):
                                continue
                            cfg = block.get("config") or {}
                            if btype == "deal_list":
                                if not cfg.get("override_urls"):
                                    games = block_games[i] if block_games and i < len(block_games) else []
                                    urls = [(g.get("link") or "").strip() for g in games if (g.get("link") or "").strip()]
                                    if urls:
                                        cfg["override_urls"] = urls
                            else:
                                if not (cfg.get("override_url") or "").strip():
                                    games = block_games[i] if block_games and i < len(block_games) else []
                                    if games and (games[0].get("link") or "").strip():
                                        cfg["override_url"] = (games[0].get("link") or "").strip()
                self._worker_busy = False
            elif kind == "error":
                _, op, e = msg
                self._worker_busy = False
                self.input_text.config(cursor="")
                if op == "load_feed":
                    messagebox.showerror("Error", f"Failed to load feed:\n{e}")
                elif op == "fetch_on_sale":
                    self.tab2_status.config(text="")
                    self._on_sale_rows = []
                    messagebox.showerror("Error", f"Failed:\n{e}")
                elif op == "email_build":
                    self.email_status_var.set("")
                    messagebox.showerror("Error", str(e))
                    import traceback
                    traceback.print_exc()
        if self._worker_busy:
            self.root.after(50, self._process_worker_queue)

    def _load_feed(self):
        if self._worker_busy:
            messagebox.showwarning("Please wait", "Another operation is in progress.")
            return
        self.input_text.config(cursor="watch")
        self._worker_busy = True
        self.root.after(50, self._process_worker_queue)

        def work():
            try:
                feed_items = fetch_and_parse()
                index = items_to_index(feed_items)
                self._worker_queue.put(("done", "load_feed", (feed_items, index)))
            except Exception as e:
                self._worker_queue.put(("error", "load_feed", e))

        threading.Thread(target=work, daemon=True).start()

    def _load_feed_tab2(self):
        if self._index is None:
            self._load_feed()
        else:
            messagebox.showinfo("Feed", "Feed already loaded.")

    def _fetch_on_sale(self):
        if self._worker_busy:
            messagebox.showwarning("Please wait", "Another operation is in progress.")
            return
        if self._index is None:
            messagebox.showwarning("Load feed first", "Load the feed from the Deal Table or Deal Finder tab first.")
            return
        self.tab2_status.config(text="Getting on-sale products...")
        self._worker_busy = True
        self.root.after(50, self._process_worker_queue)
        index = self._index

        def work():
            try:
                products = get_on_sale_products(index, resolve_steam_by_name=True)
                self._worker_queue.put(("progress", "tab2", f"Found {len(products)} on sale. Fetching Steam data..."))
                def progress(i, total):
                    self._worker_queue.put(("progress", "tab2", f"Fetching Steam data... {i}/{total}"))
                rows = enrich_with_steam_reviews(products, progress_callback=progress)
                self._worker_queue.put(("done", "fetch_on_sale", rows))
            except Exception as e:
                self._worker_queue.put(("error", "fetch_on_sale", e))

        threading.Thread(target=work, daemon=True).start()

    def _toggle_filters_tab2(self):
        """Show or hide the filter controls to free space for the results list."""
        if self._filters_visible:
            self.tab2_filt_frame.pack_forget()
            self._filters_visible = False
            self.tab2_filters_toggle_btn.config(text="Show filters \u25B6")
        else:
            self.tab2_filt_frame.pack(fill=tk.X, pady=(4, 0))
            self._filters_visible = True
            self.tab2_filters_toggle_btn.config(text="Hide filters \u25BC")

    def _apply_filters_tab2(self):
        score_type = self.score_filter_type.get()
        score_val = self.score_filter_value.get()
        label_val = self.label_filter.get()
        min_rev = self.min_reviews_var.get()
        discount_val = self.discount_filter_var.get()
        sale_end_type = self.sale_end_filter_type.get()
        sale_end_val = self.sale_end_filter_value.get()
        search_query = self.tab2_search_var.get()
        release_date_type = self.release_date_filter_type.get()
        release_date_val = self.release_date_filter_value.get()
        rows = apply_deal_filters(
            self._on_sale_rows,
            score_type=score_type,
            score_value=score_val,
            label_value=label_val,
            min_reviews=min_rev,
            discount_value=discount_val,
            sale_end_type=sale_end_type,
            sale_end_value=sale_end_val,
        )
        rows = _apply_game_search_filter(rows, search_query)
        rows = _apply_release_date_filter(rows, release_date_type, release_date_val)
        rows = _apply_publisher_filter(rows, self.tab2_publisher_filter_var.get())
        rows = _apply_developer_filter(rows, self.tab2_developer_filter_var.get())
        rows = _apply_tags_filter(rows, self.tab2_tags_filter_var.get())
        self._populate_tab2_sheet(rows)

    def _resize_tab2_columns(self, event=None):
        """Distribute column widths across the full table width."""
        try:
            w = self.tab2_sheet_frame.winfo_width()
            if w <= 1:
                return
            scrollbar_w = 20
            total = max(100, w - scrollbar_w)
            widths = [max(40, int(total * r)) for r in self._tab2_column_ratios]
            self.tab2_sheet.set_column_widths(column_widths=widths)
        except (tk.TclError, AttributeError):
            pass

    def _value_to_color(self, t: float) -> str:
        """Map t in [0, 1] to green -> yellow -> orange -> red (hex)."""
        if t <= 0:
            return "#d4edda"
        if t >= 1:
            return "#f8d7da"
        if t < 0.33:
            u = t / 0.33
            return _lerp_hex("#d4edda", "#fff3cd", u)
        if t < 0.66:
            u = (t - 0.33) / 0.33
            return _lerp_hex("#fff3cd", "#ffe5b4", u)
        u = (t - 0.66) / 0.34
        return _lerp_hex("#ffe5b4", "#f8d7da", u)

    def _populate_tab2_sheet(self, rows: list[dict]):
        self._tab2_displayed_rows = rows
        data = []
        for r in rows:
            title = (r.get("title") or "").strip() or "—"
            pct = r.get("steam_percent_positive")
            desc = r.get("steam_review_desc") or ""
            if pct is not None and desc:
                rating = f"{desc} ({pct}%)"
            elif desc:
                rating = desc
            elif pct is not None:
                rating = f"{pct}%"
            else:
                rating = "N/A"
            reviews = r.get("steam_total_reviews")
            reviews_str = str(reviews) if reviews is not None and reviews > 0 else "N/A"
            discount_str = _discount_str(r)
            release_str = _release_date_str(r)
            sale_end_str = _sale_end_str(r)
            developer_str = (r.get("steam_developer") or "").strip() or "—"
            publisher_str = (r.get("steam_publisher") or "").strip() or "—"
            tags_raw = r.get("steam_tags")
            if isinstance(tags_raw, list):
                tags_str = ", ".join(str(t).strip() for t in tags_raw if t and str(t).strip()) or "—"
            else:
                tags_str = (tags_raw or "").strip() or "—"
            data.append([title, rating, reviews_str, discount_str, release_str, sale_end_str, developer_str, publisher_str, tags_str])
        self.tab2_sheet.set_sheet_data(data)
        self._apply_tab2_color_scale(rows)
        self._resize_tab2_columns()
        self.tab2_sheet.refresh()

    def _apply_tab2_color_scale(self, rows: list[dict]):
        """Apply green->yellow->orange->red background for Rating (1), Reviews (2), % Off (3)."""
        if not rows:
            return
        n = len(rows)
        # Collect numeric values: col 1 = rating 0-100, col 2 = log10(reviews), col 3 = discount 0-100
        rating_vals = []
        review_vals = []
        discount_vals = []
        for r in rows:
            pct = r.get("steam_percent_positive")
            rating_vals.append(float(pct) if pct is not None else None)
            rev = r.get("steam_total_reviews")
            review_vals.append(math.log10(max(1, rev)) if rev and rev > 0 else None)
            discount_vals.append(float(_discount_pct(r)) if _discount_pct(r) is not None else None)
        for col_idx, vals in enumerate([rating_vals, review_vals, discount_vals]):
            numeric = [v for v in vals if v is not None]
            if not numeric:
                continue
            lo, hi = min(numeric), max(numeric)
            span = hi - lo if hi > lo else 1.0
            sheet_col = col_idx + 1
            for r in range(n):
                v = vals[r]
                if v is None:
                    continue
                t = (v - lo) / span
                color = self._value_to_color(t)
                self.tab2_sheet[(r, sheet_col)].bg = color

    def _on_tab2_sheet_double_click(self, event=None):
        """Open the selected row's product page in the default browser."""
        if not self._tab2_displayed_rows:
            return
        sel = self.tab2_sheet.get_currently_selected()
        if sel is None:
            return
        row_idx = sel.row
        if 0 <= row_idx < len(self._tab2_displayed_rows):
            link = self._tab2_displayed_rows[row_idx].get("link")
            if link:
                webbrowser.open(link)

    def _copy_tab2_urls(self):
        urls = [r.get("link") or "" for r in self._tab2_displayed_rows if r.get("link")]
        if urls:
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(urls))
            messagebox.showinfo("Copied", f"Copied {len(urls)} product URL(s) to clipboard.")
        else:
            messagebox.showwarning("Nothing to copy", "Fetch on-sale list first.")

    def _export_tab2_to_xlsx(self):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        rows = getattr(self, "_tab2_displayed_rows", None) or []
        if not rows:
            messagebox.showwarning("Nothing to export", "Fetch on-sale list and apply filters first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Deal Finder"
        headers = ["Game", "Rating", "Reviews", "% Off", "Release date", "Sale end", "Developer", "Publisher", "Tags"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)

        rating_vals, review_vals, discount_vals = [], [], []
        data_rows = []
        for r in rows:
            title = (r.get("title") or "").strip() or "—"
            pct = r.get("steam_percent_positive")
            desc = r.get("steam_review_desc") or ""
            if pct is not None and desc:
                rating = f"{desc} ({pct}%)"
            elif desc:
                rating = desc
            elif pct is not None:
                rating = f"{pct}%"
            else:
                rating = "N/A"
            reviews = r.get("steam_total_reviews")
            reviews_str = str(reviews) if reviews is not None and reviews > 0 else "N/A"
            discount_str = _discount_str(r)
            release_str = _release_date_str(r)
            sale_end_str = _sale_end_str(r)
            developer_str = (r.get("steam_developer") or "").strip() or "—"
            publisher_str = (r.get("steam_publisher") or "").strip() or "—"
            tags_raw = r.get("steam_tags")
            if isinstance(tags_raw, list):
                tags_str = ", ".join(str(t).strip() for t in tags_raw if t and str(t).strip()) or "—"
            else:
                tags_str = (tags_raw or "").strip() or "—"
            data_rows.append([title, rating, reviews_str, discount_str, release_str, sale_end_str, developer_str, publisher_str, tags_str])
            rating_vals.append(float(pct) if pct is not None else None)
            rev = r.get("steam_total_reviews")
            review_vals.append(math.log10(max(1, rev)) if rev and rev > 0 else None)
            discount_vals.append(float(_discount_pct(r)) if _discount_pct(r) is not None else None)

        for i, row_data in enumerate(data_rows):
            excel_row = i + 2
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=excel_row, column=col, value=val)
                row_fill = "FFFFFF" if (i % 2) == 0 else "F0F4F8"
                cell.fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")

        n = len(rows)
        for col_idx, vals in enumerate([rating_vals, review_vals, discount_vals]):
            numeric = [v for v in vals if v is not None]
            if not numeric:
                continue
            lo, hi = min(numeric), max(numeric)
            span = hi - lo if hi > lo else 1.0
            excel_col = col_idx + 2
            for r in range(n):
                v = vals[r]
                if v is None:
                    continue
                t = (v - lo) / span
                color = self._value_to_color(t).lstrip("#")
                ws.cell(row=r + 2, column=excel_col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to {path}")

    def _clear_steam_cache(self):
        clear_steam_cache()
        clear_app_list_cache()
        clear_steam_appdetails_cache()
        clear_name_resolution_cache()
        clear_steamspy_cache()
        messagebox.showinfo("Cache cleared", "Steam review, app list, appdetails, name-resolution, and SteamSpy caches cleared.")

    def _build_table(self):
        if self._index is None:
            self._load_feed()
        if self._index is None:
            return
        urls = parse_pasted_urls(self.input_text.get("1.0", tk.END))
        if not urls:
            messagebox.showwarning("No URLs", "Paste at least one product URL.")
            return
        products, not_found = resolve_urls_to_products(self._index, urls)
        currencies = self._get_selected_currencies()
        if not currencies:
            messagebox.showwarning("No currencies", "Select at least one currency.")
            return
        markdown = build_reddit_table(products, currencies)
        self.output_text.config(state=tk.NORMAL)
        self.output_text.delete("1.0", tk.END)
        self.output_text.insert(tk.END, markdown)
        self.output_text.config(state=tk.DISABLED)
        if not_found:
            messagebox.showwarning(
                "Some URLs not found",
                f"{len(not_found)} URL(s) were not found in the feed:\n" + "\n".join(not_found[:5])
                + ("\n..." if len(not_found) > 5 else ""),
            )
        if not products:
            messagebox.showwarning("No products", "No products could be found for the given URLs.")

    def _copy_output(self):
        text = self.output_text.get("1.0", tk.END)
        if text.strip():
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("Copied", "Reddit markdown copied to clipboard.")
        else:
            messagebox.showwarning("Nothing to copy", "Build a table first.")

    def run(self):
        self.root.mainloop()


def main():
    app = Application()
    app.run()


if __name__ == "__main__":
    main()
